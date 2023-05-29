import pyautogui
import time
import cv2
from openpyxl import load_workbook
import keyboard
import pytesseract

running = True  # Flag to indicate if the code should keep running


def stop_execution(e):
    global running
    if e.name == 'q':
        running = False


# Register the 'q' key as the trigger to stop the code
keyboard.on_press(stop_execution)

folder_path = 'C:\\Users\\user\\Desktop\\Test'
# Open Excel file and get Protocol numbers
load_excel_path = f'{folder_path}\\2019 İSTATİSTİK xlsx.xlsx'
resut_excel = f'{folder_path}\\Test.xlsx'
wbresult = load_workbook(f'{folder_path}\\Test.xlsx')
wsresult = wbresult.active
wb = load_workbook(load_excel_path)
ws = wb.active
for i in range(2, 52):
    if not running:
        print(i)
        break  # Exit the loop if 'q' key was pressed
    protokol_no = ws[f'C{i}'].value
    wsresult[f'A{i}'].value = protokol_no

    # Click to Avicenna
    pyautogui.click(30, 10)
    time.sleep(0.5)

    # Write Patient No.
    pyautogui.moveTo(100, 85)
    pyautogui.click(100, 85)
    pyautogui.click(100, 85)
    pyautogui.press('del')
    time.sleep(0.5)
    pyautogui.typewrite(str(protokol_no))
    pyautogui.press('enter')
    time.sleep(4)

    #Check if patient not found:
    screenshot1 = pyautogui.screenshot(region=(210, 0, 480, 170))
    screenshot1.save(f'{folder_path}\\screenshot1.jpg')
    image01 = cv2.imread(f'{folder_path}\\standard1.jpg')
    image02 = cv2.imread(f'{folder_path}\\screenshot1.jpg')
    image01_hsv = cv2.cvtColor(image01, cv2.COLOR_BGR2HSV)
    image02_hsv = cv2.cvtColor(image02, cv2.COLOR_BGR2HSV)
    hist01 = cv2.calcHist([image01_hsv], [0, 1], None, [180, 256], [0, 180, 0, 256])
    hist02 = cv2.calcHist([image02_hsv], [0, 1], None, [180, 256], [0, 180, 0, 256])
    correlation1 = cv2.compareHist(hist01, hist02, cv2.HISTCMP_CORREL)
    if correlation1 > 0.99:
        gray_image = cv2.cvtColor(image02, cv2.COLOR_BGR2GRAY)
        if 'ex' in pytesseract.image_to_string(gray_image):
            pyautogui.moveTo(660, 150)
            pyautogui.click(660, 150)
            time.sleep(4)
        else:
            wsresult[f'B{i}'].value = 'Yanlış'
            wbresult.save(resut_excel)
            pyautogui.moveTo(660, 150)
            pyautogui.click(660, 150)
            time.sleep(4)
            continue


    # Click to 'HASTA BAZINDA'
    pyautogui.moveTo(40, 300)
    pyautogui.click(40, 300)
    time.sleep(2)

    # Click to 'Laboratuvar'
    screenshot2 = pyautogui.screenshot(region=(100, 165, 600, 30))
    screenshot2.save(f'{folder_path}\\screenshot2.jpg')
    image2 = cv2.imread(f'{folder_path}\\screenshot2.jpg')
    gray_image = cv2.cvtColor(image2, cv2.COLOR_BGR2GRAY)
    text = pytesseract.image_to_string(gray_image)
    words = text.split()
    index = words.index('Laboratuvar')
    print(index)
    if index == 2:
        pyautogui.moveTo(190, 180)
        pyautogui.click(190, 180)
    if index == 3:
        pyautogui.moveTo(250, 180)
        pyautogui.click(250, 180)
    if index == 5:
        pyautogui.moveTo(370, 180)
        pyautogui.click(370, 180)
    if index == 7:
        pyautogui.moveTo(400, 180)
        pyautogui.click(400, 180)
    time.sleep(2)

    # Search for the test.
    pyautogui.moveTo(90, 265)
    pyautogui.click(90, 265)
    pyautogui.typewrite('menenjit')
    pyautogui.press('enter')
    time.sleep(0.5)
    screenshot = pyautogui.screenshot(region=(60, 295, 200, 390))
    screenshot.save(f'{folder_path}\\screenshot.jpg')

    image1 = cv2.imread(f'{folder_path}\\standard.jpg')
    image2 = cv2.imread(f'{folder_path}\\screenshot.jpg')

    image1_hsv = cv2.cvtColor(image1, cv2.COLOR_BGR2HSV)
    image2_hsv = cv2.cvtColor(image2, cv2.COLOR_BGR2HSV)

    hist1 = cv2.calcHist([image1_hsv], [0, 1], None, [180, 256], [0, 180, 0, 256])
    hist2 = cv2.calcHist([image2_hsv], [0, 1], None, [180, 256], [0, 180, 0, 256])

    correlation = cv2.compareHist(hist1, hist2, cv2.HISTCMP_CORREL)
    chi_square = cv2.compareHist(hist1, hist2, cv2.HISTCMP_CHISQR)
    intersection = cv2.compareHist(hist1, hist2, cv2.HISTCMP_INTERSECT)
    bhattacharyya = cv2.compareHist(hist1, hist2, cv2.HISTCMP_BHATTACHARYYA)

    if correlation == 1.0:
        wsresult[f'B{i}'].value = 0
    else:
        wsresult[f'B{i}'].value = 1
    wbresult.save(resut_excel)
